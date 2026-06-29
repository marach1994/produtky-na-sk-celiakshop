#!/usr/bin/env python3
"""
CeliakShop CZ → SK EUR feed transformer
Replikuje pravidla Mergado projektu 339108 (30 pravidel)

Použití:
  python transform.py              # stáhne CSV z URL, zapíše celiakshop_sk.csv vedle skriptu
  python transform.py input.csv    # zpracuje lokální soubor
"""

import csv
import re
import json
import sys
import os
import ssl
import urllib.request
import tempfile
import shutil
from decimal import Decimal, ROUND_HALF_UP
from io import StringIO

SOURCE_URL = (
    "https://www.celiakshop.cz/export/products.csv"
    "?patternId=61&partnerId=3"
    "&hash=95ecec03380b553c399e5a1b4a7e17d1598be40e1649a612d4cb1777adfc6429"
)

SK_EXPORT_URL = (
    "https://www.celiakshop.sk/export/products.xls"
    "?patternId=15&partnerId=3"
    "&hash=a5be98ba00c0e0a70d376a38ed95e5c7a1976789eda88449c713fee504925bbf"
)

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
CATEGORY_MAPPING_FILE = os.path.join(SCRIPT_DIR, "category_mapping_rows.json")
OUTPUT_FILE = os.path.join(SCRIPT_DIR, "celiakshop_sk.csv")

# Výstupní sloupce ve správném pořadí (dle importního formátu SK Shoptetu)
OUTPUT_COLUMNS = [
    "code", "pairCode", "supplier", "manufacturer",
    "defaultCategory", "categoryText",
    "price", "standardPrice", "purchasePrice", "actionPrice",
    "actionFrom", "actionUntil",
    "negativeAmount",
    "actionFlagActive", "actionFlagValidFrom", "actionFlagValidUntil",
    "newFlagActive", "newFlagValidFrom", "newFlagValidUntil",
    "tipFlagActive",
    "custom1FlagActive",
    "filteringProperty:Značka", "filteringProperty:Typ",
    "filteringProperty:Vlastnosti", "filteringProperty:Dodavatel",
    "custom3FlagActive", "topFlagActive", "letoFlagActive",
    "custom2FlagActive",  # = bez-lepkuFlagActive ze zdroje
    "categoryText2", "categoryText3", "categoryText4", "categoryText5",
    "categoryText6", "categoryText7", "categoryText8", "categoryText9",
    "categoryText10", "categoryText11",
]

# Pravidla 1–4: přepočet cen CZK → EUR
PRICE_RULES = {
    "actionPrice":   Decimal("24.7"),
    "price":         Decimal("24.5"),
    "standardPrice": Decimal("24.7"),
    "purchasePrice": Decimal("24.7"),
}

# Pravidlo 5: překlad vlastností CZ → SK (batch_rewriting_values, regex=false)
VLASTNOSTI_MAP = {
    "Bez přidaného cukru;Bez lepku":                                  "Bez pridaného cukru;Bez lepku",
    "Bez přidaného cukru":                                            "Bez pridaného cukru",
    "Vegan;Bez lepku":                                                "Vegan;Bez lepku",
    "Bez lepku":                                                      "Bez lepku",
    "Bez laktózy;bez palmového oleje":                                "Bez laktózy;bez palmového oleja",
    "Bez laktózy":                                                    "Bez laktózy",
    "Bez lepku;Bez laktózy":                                          "Bez lepku;Bez laktózy",
    "Vegan;Bez lepku;Bez laktózy":                                    "Vegan;Bez lepku;Bez laktózy",
    "Vegan;Bez lepku;Bio":                                            "Vegan;Bez lepku;Bio",
    "Bez lepku;Bio":                                                  "Bez lepku;Bio",
    "Bez přidaného cukru;Vegan;Bez laktózy;bez palmového oleje":      "Bez pridaného cukru;Vegan;Bez laktózy;bez palmového oleja",
    "Bez přidaného cukru;Vegan;Bez laktózy;Bio;bez palmového oleje":  "Bez pridaného cukru;Vegan;Bez laktózy;Bio;bez palmového oleja",
    "Bez přidaného cukru;Bez lepku;Bio;bez palmového oleje":          "Bez pridaného cukru;Bez lepku;Bio;bez palmového oleja",
    "bez palmového oleje":                                            "bez palmového oleja",
    "Bez lepku;bez palmového oleje":                                  "Bez lepku;bez palmového oleja",
    "Bez přidaného cukru;Vegan;Bez lepku;Bio":                       "Bez pridaného cukru;Vegan;Bez lepku;Bio",
    "Bio":                                                            "Bio",
    "Bez přidaného cukru;Bez lepku;Bio":                             "Bez pridaného cukru;Bez lepku;Bio",
    "Bez přidaného cukru;Bez lepku;Bez laktózy":                     "Bez pridaného cukru;Bez lepku;Bez laktózy",
}

# Pole kategorií (pravidla 7–30)
CATEGORY_FIELDS = [
    "defaultCategory",
    "categoryText", "categoryText2", "categoryText3", "categoryText4",
    "categoryText5", "categoryText6", "categoryText7", "categoryText8",
    "categoryText9", "categoryText10", "categoryText11",
]


def mathematical_round(value: Decimal, places: int) -> Decimal:
    quantizer = Decimal(10) ** -places
    return value.quantize(quantizer, rounding=ROUND_HALF_UP)


def convert_price(value_str: str, divisor: Decimal) -> str:
    if not value_str:
        return value_str
    try:
        val = Decimal(value_str.replace(",", "."))
        result = mathematical_round(val / divisor, 2)
        return str(result).replace(".", ",")
    except Exception:
        return value_str


def apply_replacing_rules(value: str) -> str:
    """Pravidla 7–18: odstraní '^>> ' a nahradí '>>' za '>'."""
    if not value:
        return value
    value = re.sub(r"^>> ", "", value, flags=re.IGNORECASE)
    value = value.replace(">>", ">")
    return value


def load_category_mapping() -> list:
    with open(CATEGORY_MAPPING_FILE, encoding="utf-8") as f:
        rows = json.load(f)
    return [
        (re.compile(row["input_value"], re.IGNORECASE), row["output_value"])
        for row in sorted(rows, key=lambda r: r["position"])
    ]


def apply_category_mapping(value: str, patterns: list) -> str:
    """Pravidla 19–30: mapuje celou hodnotu, vrátí první shodu."""
    if not value:
        return value
    for compiled_re, output in patterns:
        if compiled_re.fullmatch(value):
            return output
    return value


def transform_row(src: dict, category_patterns: list) -> dict:
    row = dict(src)

    # Pravidla 1–4: přepočet cen
    for field, divisor in PRICE_RULES.items():
        if row.get(field):
            row[field] = convert_price(row[field], divisor)

    # Pravidlo 5: Vlastnosti
    prop = "filteringProperty:Vlastnosti"
    if row.get(prop) in VLASTNOSTI_MAP:
        row[prop] = VLASTNOSTI_MAP[row[prop]]

    # Pravidla 7–30: transformace kategorií
    for cat_field in CATEGORY_FIELDS:
        if row.get(cat_field):
            val = apply_replacing_rules(row[cat_field])
            val = apply_category_mapping(val, category_patterns)
            row[cat_field] = val

    # Mapování sloupce: bez-lepkuFlagActive → custom2FlagActive
    row["custom2FlagActive"] = row.get("bez-lepkuFlagActive", "")

    # Vrátí pouze výstupní sloupce ve správném pořadí
    return {col: row.get(col, "") for col in OUTPUT_COLUMNS}


def fetch_url(url: str) -> bytes:
    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE
    with urllib.request.urlopen(url, context=ctx) as resp:
        return resp.read()


def load_sk_codes() -> set:
    """Stáhne SK export a vrátí sadu existujících kódů produktů."""
    import openpyxl
    print("Stahuji SK produkty pro filtrování...", file=sys.stderr)
    data = fetch_url(SK_EXPORT_URL)
    # SK export je xlsx přejmenovaný na xls — openpyxl ho zvládne po uložení
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    try:
        tmp.write(data)
        tmp.close()
        wb = openpyxl.load_workbook(tmp.name, read_only=True)
        ws = wb.active
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        code_idx = headers.index("code")
        codes = {str(row[code_idx]) for row in ws.iter_rows(min_row=2, values_only=True) if row[code_idx]}
        wb.close()
    finally:
        os.unlink(tmp.name)
    print(f"SK eshop: {len(codes)} produktů nalezeno.", file=sys.stderr)
    return codes


def main():
    category_patterns = load_category_mapping()

    if len(sys.argv) >= 2:
        src_path = sys.argv[1]
        with open(src_path, encoding="utf-8-sig") as f:
            content = f.read()
    else:
        print(f"Stahuji CZ feed...", file=sys.stderr)
        content = fetch_url(SOURCE_URL).decode("utf-8-sig")

    sk_codes = load_sk_codes()

    reader = csv.DictReader(StringIO(content), delimiter=";")

    with open(OUTPUT_FILE, "w", encoding="utf-8-sig", newline="") as out:
        writer = csv.DictWriter(
            out, fieldnames=OUTPUT_COLUMNS, delimiter=";",
            quoting=csv.QUOTE_MINIMAL, lineterminator="\r\n",
        )
        writer.writeheader()

        count = 0
        missing = []
        for row in reader:
            if row.get("code") not in sk_codes:
                missing.append((row.get("code", ""), row.get("name", "")))
                continue
            writer.writerow(transform_row(row, category_patterns))
            count += 1

    print(f"\nHotovo: {count} produktů zapsáno → {OUTPUT_FILE}", file=sys.stderr)
    if missing:
        print(f"\nProdukty z CZ feedu které NEJSOU na SK ({len(missing)} ks):", file=sys.stderr)
        for code, name in sorted(missing):
            print(f"  {code:<15} {name}", file=sys.stderr)


if __name__ == "__main__":
    main()
